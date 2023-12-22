from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import json


print()
print('Grabbing report data...')
print()


# Function that defines columns based on header titles
def defineColumn(headerSheetName, headers, headerrow):
    for col in range(1, headerSheetName.max_column + 1):
        for header in headers:
            if headerSheetName[get_column_letter(col) + str(headerrow)].value != None:
                if headerSheetName[get_column_letter(col) + str(headerrow)].value.upper().lstrip().rstrip() == header.upper():
                    return get_column_letter(col)


def getSkillColLetter(tier, num):
    if tier == 'Move':
        if num == 1:
            return 'D'
        elif num == 2:
            return 'E'
        elif num == 3:
            return 'F'
        elif num == 4:
            return 'G'
        elif num == 5:
            return 'H'
    elif tier == 'Develop':
        if num == 1:
            return 'I'
        elif num == 2:
            return 'J'
        elif num == 3:
            return 'K'
        elif num == 4:
            return 'L'
        elif num == 5:
            return 'M'
    elif tier == 'Connect':
        if num == 1:
            return 'N'
        elif num == 2:
            return 'O'
        elif num == 3:
            return 'P'
        elif num == 4:
            return 'Q'
        elif num == 5:
            return 'R'


def getSkillColumn(skill):
    if skill == '3':
        return 'D'
    elif skill == '2':
        return 'E'
    elif skill == '1':
        return 'F'


try:
    # Classes workbook
    data_wb = load_workbook(filename='inputs/Check-In Cards (Data File).xlsx')
except:
    input('Error')

try:
    # Template workbook
    template_wb = load_workbook(
        filename='templates/Check-In Cards (template).xlsx')
    template = template_wb.active
except:
    input('Error 2')

students_list = []

# iterate through teacher sheets
for teacher in data_wb.sheetnames:

    _data = data_wb[teacher]

    for row in range(2, _data.max_row + 1):

        try:

            # Vars
            _tier = _data['C'+str(row)].value

            _skill_1 = _data[getSkillColLetter(
                _tier, 1)+'1'].value + '***' + str(_data[getSkillColLetter(_tier, 1)+str(row)].value)
            _skill_2 = _data[getSkillColLetter(
                _tier, 2)+'1'].value + '***' + str(_data[getSkillColLetter(_tier, 2)+str(row)].value)
            _skill_3 = _data[getSkillColLetter(
                _tier, 3)+'1'].value + '***' + str(_data[getSkillColLetter(_tier, 3)+str(row)].value)
            _skill_4 = _data[getSkillColLetter(
                _tier, 4)+'1'].value + '***' + str(_data[getSkillColLetter(_tier, 4)+str(row)].value)
            _skill_5 = _data[getSkillColLetter(
                _tier, 5)+'1'].value + '***' + str(_data[getSkillColLetter(_tier, 5)+str(row)].value)

            _card = {
                "teacher": teacher,
                "skill_1": _skill_1,
                "skill_2": _skill_2,
                "skill_3": _skill_3,
                "skill_4": _skill_4,
                "skill_5": _skill_5,
                "comments": _data['S'+str(row)].value
            }

            # Student's first card
            if len(list(filter(lambda s: s['name'] == _data['A'+str(row)].value, students_list))) == 0:
                students_list.append({
                    "name": _data['A'+str(row)].value,
                    "tier": _data['C'+str(row)].value,
                    "cards": [_card]
                })

            # Add card if student's already has a card
            else:
                for student in students_list:
                    if student['name'] == _data['A'+str(row)].value:
                        student["cards"].append(_card)
                        break

        except Exception as e:
            print(e)
            print(teacher)
            print(_data['A'+str(row)].value)

# Populate check in cards with data from students_list
for student in sorted(students_list, key=lambda s: s['name']):

    num = 1

    for card in student['cards']:

        checkin = template_wb.copy_worksheet(template)
        checkin.title = str(num) + ' ' + \
            student['name'] + ' - ' + card['teacher']

        num += 1

        img = Image('utils/logo.png')
        img.height = 120
        img.width = 220
        checkin.add_image(img, 'C1')

        # Heading
        if student['tier'] == 'Move':
            checkin['A8'] = 'Move -- Mini-Movers / Newbies / Petites'
            checkin['A9'] = 'explore | create | belong'

        if student['tier'] == 'Develop':
            checkin['A8'] = 'Move -- Minis / Beginners'
            checkin['A9'] = 'empower | friendship | evolve'

        if student['tier'] == 'Connect':
            checkin['A8'] = 'Connect -- Intermediate / Advanced'
            checkin['A9'] = 'celebrate | lead | inpsire'

        # Name and instructor
        checkin['A10'] = student['name']
        checkin['A12'] = card['teacher']

        # Skills
        checkin['B14'] = card['skill_1'].split('***')[0]

        if getSkillColumn(card['skill_1'].split('***')[1]) != None:
            checkin[getSkillColumn(
                card['skill_1'].split('***')[1]) + '14'] = '✓'

        if getSkillColumn(card['skill_2'].split('***')[1]) != None:
            checkin['B15'] = card['skill_2'].split('***')[0]
            checkin[getSkillColumn(
                card['skill_2'].split('***')[1]) + '15'] = '✓'

        if getSkillColumn(card['skill_3'].split('***')[1]) != None:
            checkin['B16'] = card['skill_3'].split('***')[0]
            checkin[getSkillColumn(
                card['skill_3'].split('***')[1]) + '16'] = '✓'

        if getSkillColumn(card['skill_4'].split('***')[1]) != None:
            checkin['B17'] = card['skill_4'].split('***')[0]
            checkin[getSkillColumn(
                card['skill_4'].split('***')[1]) + '17'] = '✓'

        if getSkillColumn(card['skill_5'].split('***')[1]) != None:
            checkin['B18'] = card['skill_5'].split('***')[0]
            checkin[getSkillColumn(
                card['skill_5'].split('***')[1]) + '18'] = '✓'

        checkin['A20'] = card['comments']

    print(student['name'] + ' check-ins created.')

del template_wb['template']

# ###### ###### ##    ## ######
# ##     ##  ## ##    ## ##
# ###### ###### ##    ## #####
#     ## ##  ##  ##  ##  ##
# ###### ##  ##    ##    ######

print()
print('Saving workbook...')
# # create json/csv file
# jsonfile = open("students.json", 'a')
# jsonfile.truncate(0)
# jsonfile.write(json.dumps(students_list, indent=4))
# jsonfile.close()

template_wb.save('Check-In Cards.xlsx')

print()
input('Complete! Press any key to exit')
print()
