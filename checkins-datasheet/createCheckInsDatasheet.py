from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Function that defines columns based on header titles


def defineColumn(headerSheetName, headers, headerrow):
    for col in range(1, headerSheetName.max_column + 1):
        for header in headers:
            if headerSheetName[get_column_letter(col) + str(headerrow)].value != None:
                if headerSheetName[get_column_letter(col) + str(headerrow)].value.upper().lstrip().rstrip() == header.upper():
                    return get_column_letter(col)


# Classes workbook
try:
    classes_wb = load_workbook(
        filename='inputs/EnrollmentDetailRpt.xlsx')
    classes = classes_wb.active
except:
    print()
    input('Incorrect file name in inputs folder. Please make sure the file is named "EnrollmentDetailRpt.xlsx".')

# Check in workbook
check_in_wb = Workbook()

# Classes columns
teachers_col = defineColumn(classes, ['Instructors'], 1)
level_col = defineColumn(classes, ['Cat1'], 1)
class_col = defineColumn(classes, ['Class Name'], 1)
first_name_col = defineColumn(classes, ['Student First Name'], 1)
last_name_col = defineColumn(classes, ['Student Last Name'], 1)

# Create list of teachers, students and levels
teachers_list = []

print()
print('Reading class file from jackrabbit and builing the data input file...')

for row in range(2, classes.max_row+1):

    _level = classes[level_col+str(row)].value
    _teachers = classes[teachers_col+str(row)].value

    if (_level not in ['Adults', 'Company', 'misc']):

        for teacher in _teachers.split(', '):
            if len(list(filter(lambda t: t['name'] == teacher, teachers_list))) == 0:
                teachers_list.append({
                    "name": teacher,
                    "students": []
                })

# Add students to teacher
for teacher in teachers_list:

    for row in range(2, classes.max_row+1):

        _level = classes[level_col+str(row)].value
        _teachers = classes[teachers_col+str(row)].value
        _first_name = classes[first_name_col+str(row)].value
        _last_name = classes[last_name_col+str(row)].value

        if _level not in ['Adults', 'Company', 'misc'] and _teachers != None and teacher["name"] in _teachers and len(list(filter(lambda s: s['name'] == _last_name + ', ' + _first_name, teacher['students']))) == 0:

            _tier = ''

            if _level in ['MiniMovers', 'Newbies', 'Petites']:
                _tier = 'Move'

            if _level in ['Minis', 'Beginners']:
                _tier = 'Develop'

            if _level in ['Intermediate', 'Advanced', 'Intermediate/Advanced']:
                _tier = 'Connect'

            teacher["students"].append({
                "name": _last_name + ', ' + _first_name,
                "level": _level,
                "tier": _tier
            })

# Create check in card data sheet using teacher objects
for teacher in teachers_list:
    _ws = check_in_wb.create_sheet(teacher["name"])

    _ws['A1'] = "Student"
    _ws['B1'] = "Level"
    _ws['C1'] = "Tier"

    # Move
    _ws['D1'] = 'Listens to Directions'
    _ws['E1'] = 'Stays on Spot'
    _ws['F1'] = 'Class Participation'
    _ws['G1'] = 'Retention of Skills & Steps'
    _ws['H1'] = 'Overall Behavior'

    # Develop
    _ws['I1'] = 'Focus in Class'
    _ws['J1'] = 'Body Control/Alignment'
    _ws['K1'] = 'Flexability'
    _ws['L1'] = 'Retention of Skills & Steps'
    _ws['M1'] = 'Overall Behavior'

    # Connect
    _ws['N1'] = 'Focus in Class'
    _ws['O1'] = 'Body Control/Alignment'
    _ws['P1'] = 'Flexability'
    _ws['Q1'] = 'Ability to Pick Up / Remember Choreography'
    _ws['R1'] = 'Respectfulness'

    # Additional comments
    _ws['S1'] = 'Additional Comments'

    for student in teacher['students']:
        _ws['A'+str(_ws.max_row+1)] = student['name']
        _ws['B'+str(_ws.max_row)] = student['level']
        _ws['C'+str(_ws.max_row)] = student['tier']

        # Move
        if student['tier'] == 'Move':
            for l in ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
                _ws[l+str(_ws.max_row)].fill = PatternFill(start_color='f8bdce',
                                                           end_color='f8bdce',
                                                           fill_type='solid')
        # Develop
        if student['tier'] == 'Develop':
            for l in ['D', 'E', 'F', 'G', 'H', 'N', 'O', 'P', 'Q', 'R']:
                _ws[l+str(_ws.max_row)].fill = PatternFill(start_color='f8bdce',
                                                           end_color='f8bdce',
                                                           fill_type='solid')
        # Connect
        if student['tier'] == 'Connect':
            for l in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                _ws[l+str(_ws.max_row)].fill = PatternFill(start_color='f8bdce',
                                                           end_color='f8bdce',
                                                           fill_type='solid')

del check_in_wb['Sheet']

# ###### ###### ##    ## ######
# ##     ##  ## ##    ## ##
# ###### ###### ##    ## #####
#     ## ##  ##  ##  ##  ##
# ###### ##  ##    ##    ######

check_in_wb.save('Check-In Cards (Data File).xlsx')

print()
input('Complete! Press any key to exit')
print()
